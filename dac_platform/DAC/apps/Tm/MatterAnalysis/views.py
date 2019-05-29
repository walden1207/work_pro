#!/usr/bin/python
# coding:utf-8
from django.db.models import Q
from django.views.decorators.http import require_POST
import os
import xlrd
import xlwt
import openpyxl

from apps.Tm.MatterAnalysis.models import DeviceTemplates, ProtocolDNC, ProtocolType, ProtocolPLC, DeviceModal, \
    Variables, ProtocolParamsDNC
from apps.Tm.GateAccess.models import GateInfoModel
from apps.Public import restful
from apps.Tm.MatterAnalysis.config import plc_protocol
from apps.User.models import UserModel
from DAC.settings import BASE_DIR
from .config import protocol_plc_list

# 暂存placeholder
placeholder = []

@require_POST
def new_device(request):
    """
    新建物影子
    :return:
    """
    token = request.POST["token"]
    gateMac = request.POST["gateMac"]
    gateType = request.POST["gateType"]
    deviceName = request.POST["deviceName"]
    deviceCode = request.POST["deviceCode"]
    des = request.POST["deviceDesc"]
    selectParams = eval(request.POST["selectedProtocol"])
    configs = eval(request.POST["config"])
    compID = UserModel.objects.filter(token=token)[0].compID
    deviceData = {
        "modalName": deviceName,
        "machCode": deviceCode,
        "des": des,
        "gateMac": gateMac,
        "protocol": ",".join(selectParams),
        "operationFlag": 9
    }
    # 建物影子之前判断是否已经存在该设备
    device = DeviceModal.objects.filter(machCode=deviceCode)
    if not device:
        DeviceModal.objects.create(**deviceData)
    else:
        device.update(**deviceData)
        Variables.objects.filter(machCode=deviceCode).delete()
        if gateType == "GW_PLC":
            ProtocolPLC.objects.filter(machCode=deviceCode).delete()
        else:
            ProtocolDNC.objects.filter(machCode=deviceCode).delete()
    # 遍历所有协议数据数据
    for i, values in configs.items():
        # 将每个协议对应的所有需要保存的数据，以对象的形式存储到列表中，进行单协议批量存储，减少数据库的save操作
        data_object_list_plc = []
        data_object_list_dnc = []
        params_object_list_plc = []
        params_object_list_dnc = []
        # data和params用来存储每一条数据，通过data和params去生成一个ORM对象
        # 存储进解析表的数据
        data = {
            "gateMac": gateMac,
            "machCode": deviceCode,
            "deviceType": ProtocolType.objects.get(TypeList=i)
        }
        # 存储进variable表的数据
        params = {
            "compID": compID,
            "machID": DeviceModal.objects.get(machCode=deviceCode),
            "machCode": deviceCode,
            "gateMac": gateMac,
            "Protocol": i
        }
        # PLC数据存储
        if gateType == "GW_PLC":
            params["operationFlag"] = 9
            for j in range(len(values)):
                params["operationValue"] = values[j][0]["value"]
                params["dataType"] = values[j][1]["selectedItem"]
                params["operationUnit"] = values[j][6]["value"]
                if i in protocol_plc_list:
                    params["gateWayIP"] = values[j][7]["value"]
                else:
                    params["serial_num"] = values[j][7]["value"]
                    params["gateWayID"] = values[j][8]["value"]
                params_object_list_plc.append(Variables(**params))

                data["name"] = values[j][0]["value"]
                data["type"] = values[j][1]["selectedItem"]
                data["functionCode"] = values[j][2]["selectedItem"]
                data["startAddress"] = values[j][3]["value"]
                data["len"] = values[j][4]["value"]
                data["delayMs"] = values[j][5]["value"]
                data["operationUnit"] = values[j][6]["value"]
                if i in protocol_plc_list:
                    data["deviceIP"] = values[j][7]["value"]
                else:
                    data["serialNum"] = values[j][7]["value"]
                    data["deviceID"] = values[j][8]["value"]
                data_object_list_plc.append(ProtocolPLC(**data))
            try:
                Variables.objects.bulk_create(params_object_list_plc)
                ProtocolPLC.objects.bulk_create(data_object_list_plc)
            except Exception as e:
                continue
        # DNC数据存储
        else:
            for j in range(0, len(values["normal"])):
                if values["normal"] and values["normal"][j]["upload"] == 9:
                    params["operationValue"] = values["normal"][j]["name"]
                    params["operationUnit"] = values["normal"][j]["unit"]
                    params["operationFlag"] = values["normal"][j]["upload"]
                    params_object_list_dnc.append(Variables(**params))

                if values["normal"]:
                    data["operationValue"] = values["normal"][j]["name"]
                    data["operationUnit"] = values["normal"][j]["unit"]
                    data["upload"] = values["normal"][j]["upload"]
                    data["Description"] = values["normal"][j]["description"]
                    data["configValue"] = values["normal"][j]["param"]
                    data["type"] = \
                        ProtocolParamsDNC.objects.filter(operationValue=data["operationValue"], ProType=i)[0].type
                if values["advanced"]:
                    advancedlist = [i["name"] for i in values["advanced"]]
                    com = ''
                    for k in values["advanced"]:
                        if k["name"] == "IP地址":
                            data["deviceIP"] = k["value"]

                        if k["name"] == "CNC端口":
                            com = k["value"]
                        if k["name"] == "PLC端口":
                            com = com + "-" + k["value"]
                    if com:
                        DeviceModal.objects.filter(machCode=deviceCode).update(serialPort=com)
                data_object_list_dnc.append(ProtocolDNC(**data))
            try:
                Variables.objects.bulk_create(params_object_list_dnc)
                ProtocolDNC.objects.bulk_create(data_object_list_dnc)
            except Exception as e:
                continue
    return restful.ok()


@require_POST
def get_protocols(request):
    """
    根据DNC\PLC返回协议列表
    :return:
    """
    # 获取协议类型
    proType = request.POST["gateType"]
    infos = ProtocolType.objects.filter(DeviceType=proType, operationFlag=9)

    if proType == "GW_DNC":
        data = {}
        typelist = []
        pro_versions = [info.TypeList.split("_") for info in infos]
        for pro in pro_versions:
            # 已有改协议，则添加版本信息
            if pro[0] in typelist:
                if len(pro) == 1:
                    data[pro[0]].append("")
                else:
                    data[pro[0]].append(pro[1])
            else:
                typelist.append(pro[0])
                data[pro[0]] = [pro[1]] if (len(pro) == 2) else []

    # 返回PLC协议信息
    else:
        data = [j for j in set(i.TypeList for i in infos)]
    return restful.ok(data)


@require_POST
def get_templateList(request):
    """
    获取GW_DNC/GW_PLC模板列表
    :return:
    """
    proType = request.POST["gateType"]
    templates = DeviceTemplates.objects.filter(proType=proType, operationFlag=9)
    data = []
    for temp in templates:
        info = {
            "templateID": temp.id,
            "name": temp.TemplateName,
            "protocol": [pro for pro in temp.TypeList.split(",")],
            "createTime": temp.createTime.strftime("%Y-%m-%d %H:%M:%S"),
        }
        data.append(info)
    return restful.ok(data)


@require_POST
def get_params_infos(request):
    """
    获取协议模板采集项详情
    :return:
    """
    # 获取网关协议的类型(PLC/DNC)
    gateType = request.POST["gateType"]
    gateMac = request.POST["gateMac"]
    machCode = request.POST["deviceCode"]
    # 获取采集项的方式
    type = request.POST["type"]
    # 按照协议获取采集项详情
    # data = []
    # 通过模板
    if type == "template":
        # 获取模板名称
        name = request.POST["name"]
        machCode = None
        templateinfo = DeviceTemplates.objects.filter(TemplateName=name, operationFlag=9)[0]
        typelist = templateinfo.TypeList.split(",")
        templateID = templateinfo.id
        data = get_params(gateMac, gateType, typelist, machCode, True, templateID)
    # 通过协议type=protocol
    else:
        # 已设置的协议
        oldProtocols = eval(request.POST["oldProtocols"])
        # 需删除的协议
        deleteProtocols = eval(request.POST["deleteProtocols"])
        # 剔除已删掉的协议
        if deleteProtocols:
            for pro in deleteProtocols:
                oldProtocols.pop(oldProtocols.index(pro))
        old_data = get_params(gateMac, gateType, oldProtocols, machCode, True)

        # 新添加协议
        newProtocols = eval(request.POST["newProtocols"])
        new_data = get_params(gateMac, gateType, newProtocols, machCode)
        data = dict(old_data, **new_data)
    return restful.ok(data)


def get_params(gateMac, gateType, typelist, machCode, flag=False, template=-1):
    """
    # 获取协议相关采集项详细信息
    :param gateMac: 网关Mac
    :param gateType: 网关协议类型
    :param typelist: 协议列表
    :param machCode: 设备编号
    :param flag: GW_DNC{
                 "True":从ProtocolDNC获取采集项信息
                 "False":从ProtocolParamsDNC获取采集项信息
                 }
    :param template: 模板ID，不是通过模板创建的template=-1
    :return:
    """
    data = {}
    global placeholder
    if not placeholder:
        datas = ProtocolParamsDNC.objects.all()
        placeholder = [j.__dict__ for j in datas]
    # 获取协议相关采集项详细信息
    if gateType == "GW_DNC":
        for type in typelist:
            # 构造返回数据的结构
            data[type] = {
                "normal": [],
                "advanced": [
                ]
            }
            # IP
            IPconfig = {
                "name": "IP地址",
                "value": "",
                "upload": 9
            }

            if flag:

                deviceType = ProtocolType.objects.get(TypeList=type).id
                param = {
                    "deviceType": deviceType,
                    "operationFlag": 9,
                    "template": template
                }
                if machCode:
                    param["machCode"] = machCode
                    param["gateMac"] = gateMac
                infos = ProtocolDNC.objects.filter(**param)
                IPconfig["value"] = infos[0].deviceIP if infos and infos[0].deviceIP and machCode else ""
            else:
                infos = ProtocolParamsDNC.objects.filter(ProType=type, operationFlag=9)
            # 添加高级选项
            data[type]["advanced"].append(IPconfig)

            for info in infos:
                # 添加基本项
                normal = {
                    "name": info.operationValue if info.operationValue else "",
                    "unit": info.operationUnit if info.operationUnit else "",
                    "upload": info.upload,
                    "description": info.Description if info.Description else "",
                    "disabled":   1,
                }
                # 处理placeholder
                if flag:
                    normal["param"] = info.configValue if info.configValue else ""
                    normal["placeholder"] = list(filter(
                        lambda x: x["operationValue"] == info.operationValue and
                                  x["ProType"] == info.deviceType.TypeList, placeholder))[0]["configValue"]
                else:
                    normal["param"] = ""
                    normal["placeholder"] = list(filter(
                        lambda x: x["operationValue"] == info.operationValue and
                                  x["ProType"] == info.ProType, placeholder))[0]["configValue"]
                data[type]["normal"].append(normal)
                if normal["placeholder"]:
                    normal["disabled"] = 0
            # 返回高级选项
            spec_protype = ["ABB", "Hanslaser", "OPCDA"]
            if type.split("_")[0] in spec_protype:
                if DeviceModal.objects.filter(machCode=machCode).count():
                    port = DeviceModal.objects.filter(machCode=machCode)[0].serialPort
                else:
                    port = ""
                advanced = {
                    "name": "CNC端口",
                    "value": port if port else "",
                    "upload": 9
                }
                if port:
                    advanced["value"] = port.split("-")[0]
                data[type]["advanced"].append(advanced)
                if type == "Hanslaser":
                    advanced = {
                        "name": "PLC端口",
                        "value": "",
                        "upload": 9
                    }
                    if port and len(port.split("-")) == 2:
                        advanced["value"] = port.split("-")[1]
                    data[type]["advanced"].append(advanced)

    # PLC模板采集项
    else:
        basic = {}
        if typelist:
            for type in typelist:
                if flag:
                    data2 = []
                    params = {
                        "deviceType": type,
                        "operationFlag": 9,
                        "template": template
                    }
                    if machCode:
                        params["machCode"] = machCode
                        params["gateMac"] = gateMac
                    # 构造返回数据的结构
                    # 有模板id,从PLC协议解析表模型拿数据
                    # deviceType = ProtocolType.objects.get(TypeList=type)
                    infos = ProtocolPLC.objects.filter(**params)
                    for info in infos:
                        data1 = plc_protocol[type][0]
                        data1[0]["value"] = info.name
                        data1[1]["selectedItem"] = info.type
                        data1[2]["selectedItem"] = info.functionCode
                        data1[3]["value"] = info.startAddress
                        data1[4]["value"] = info.len
                        data1[5]["value"] = info.delayMs
                        data1[6]["value"] = info.operationUnit
                        if type in protocol_plc_list:
                            data1[7]["value"] = info.deviceIP
                        else:
                            data1[7]["value"] = info.serialNum
                            data1[8]["value"] = info.deviceID
                        data2.append(data1)
                        basic[type] = data2
                else:
                    infos = plc_protocol[type]
                    basic[type] = infos
            data = basic

    return data


def del_params(gateType, gateMac, deviceCode, typeList):
    """
    删除相关协议的采集项信息
    :param gateType:
    :param gateMac:
    :param deviceCode:
    :param typeList:
    :return:
    """
    if gateType == "GW_DNC":
        ProtocolDNC.objects.filter(gateMac=gateMac, deviceType__in=typeList).update(operationFlag=0)


@require_POST
def get_MatterAnalysis(request):
    """
    获取物解析列表
    :return:
    """
    token = request.POST["token"]
    # 当前查看页
    page = int(request.POST["page"])
    # 每页显示的条数
    limit = int(request.POST["limit"])
    # 搜索关键字, ""表示查询所有
    searchKey = request.POST["searchKey"]
    # 升降序
    order = request.POST["order"]
    # 排序字段
    orderProp = request.POST["orderProp"]
    # 取每一页的第一行和最后一行数的下标
    start = (page - 1) * limit
    end = start + limit
    # 存储查询出的网关数据
    gate_list = []
    user = UserModel.objects.filter(token=token).all()
    if user:
        compID = user[0].compID
    else:
        compID = ""
    if compID:
        # 管理员
        if compID == -1:
            # 所有网关
            gateInfos = GateInfoModel.objects.filter(ifregister=1)
        else:
            # 根据公司来查询
            compID = str(compID).zfill(4)
            # 该公司所有网关
            gateInfos = GateInfoModel.objects.filter(compID=compID, ifregister=1)
        # 根据搜索关键字来查询
        if not searchKey:
            gateInfo1 = gateInfos.filter()
        else:
            # 查询出关联的设备编码和协议类型
            devices = DeviceModal.objects.filter(gateMac__in=[i.gateMac for i in gateInfos]).filter(
                Q(protocol__icontains=searchKey) | Q(machCode__icontains=searchKey))
            # 根据搜索字来查询，当搜索协议或者machCode时，根据gateMac__in可以查询出有这些协议或者machCode的网关
            gateInfo1 = gateInfos.filter(Q(gateName__icontains=searchKey) |
                                         Q(gateMac__icontains=searchKey) |
                                         Q(gateType__icontains=searchKey) |
                                         Q(machID__icontains=searchKey) |
                                         Q(gateMac__in=[i.gateMac for i in devices]))
        # 对数据进行排序
        if order == "ascending":
            gateInfo1 = gateInfo1.order_by(orderProp)
        else:
            gateInfo1 = gateInfo1.order_by("-" + orderProp)
        # 总的网关数
        totalCount = len(gateInfo1)
        # 拿出分页数据
        gateInfo2 = gateInfo1[start:end]
        # 遍历查询出的数据，存储到gate_list中，以便返回给前端进行展示
        for index, i in enumerate(gateInfo2):
            protocol = []
            data = {
                "id": i.id,
                "gateName": i.gateName,
                "compID": i.compID,
                "gateType": i.gateType,
                "gateMac": i.gateMac,
                "serverIP": i.ServerIP,
                "status": i.status,
                "software_version": i.software_version,
                # 网关的操作状态
                "operateStatus": ""
            }
            devices = DeviceModal.objects.filter(gateMac=i.gateMac)
            # 获取网关对应的machCode
            data["machCode"] = [x.machCode for x in devices]
            # 获取网关对应的协议，协议是与物影子进行绑定的，需要通过物影子获取
            for x in devices:
                device_protocol = DeviceModal.objects.filter(machCode=x.machCode)[0].protocol
                if device_protocol:
                    protocol.extend([i for i in device_protocol.split(",")])
            protocol = list(set(protocol))
            data["protocols"] = protocol
            gate_list.append(data)
        gate_infos = {"items": gate_list, "totalCount": totalCount}
        result = restful.ok(data=gate_infos)
    else:
        result = restful.params_error(message="token is worng!")
    return result


@require_POST
def get_deviceList(request):
    """
    获取物影子列表
    :return:
    """
    gateMac = request.POST["gateMac"]
    data = []
    # 查询物影子表
    infos = DeviceModal.objects.filter(gateMac=gateMac, operationFlag=9)
    for info in infos:
        basic = {
            "deviceName": info.modalName,
            "deviceCode": info.machCode,
            "deviceDesc": info.des,
            "protocols": [],
        }
        # 加入未对选择协议情况的异常处理
        if info.protocol:
            basic["protocols"] = [i for i in info.protocol.split(",")]
        data.append(basic)
    return restful.ok(data)


@require_POST
def get_dnc_params(request):
    """
    根据DNC协议返回采集项
    :return:
    """
    type = request.POST["type"]
    infos = ProtocolDNC.objects.filter(type_in=type).all()

    return restful.ok()


@require_POST
def delete_device(request):
    """
    删除物影子
    :return:
    """
    deviceCode = request.POST["deviceCode"]
    gateMac = request.POST["gateMac"]
    gateType = request.POST["gateType"]
    try:
        Variables.objects.filter(machCode=deviceCode).delete()
        if gateType == "GW_DNC":
            ProtocolDNC.objects.filter(machCode=deviceCode).delete()
        else:
            ProtocolPLC.objects.filter(machCode=deviceCode).delete()
        # 将deviceMoal表中的数据清除
        DeviceModal.objects.filter(machCode=deviceCode).delete()
        result = restful.ok()
    except Exception as e:
        result = restful.params_error(message="删除网关失败!")
    return result


@require_POST
def get_plc_params(request):
    """
    获取plc采集项
    :return:
    """
    data = {}
    protocols = request.POST["protocols"].split(",")
    if protocols:
        for i in protocols:
            if i in plc_protocol:
                data[i] = plc_protocol[i]
        result = restful.ok(data=data)
    else:
        result = restful.params_error(message="param is null!")
    return result


@require_POST
def new_template(request):
    """
    新建协议模板
    :return:
    """
    # 获取POST数据
    gateType = request.POST["gateType"]
    selectParams = eval(request.POST["selectedProtocol"])

    """
    下面的代码块是保存模板
    """
    params = {
        "TemplateName": request.POST['name'],
        "proType": gateType,
        "TypeList": ",".join(selectParams),
    }
    count = DeviceTemplates.objects.filter(TemplateName=params["TemplateName"]).count()
    if count:
        return restful.params_error(message="模板名重复")
    else:
        DeviceTemplates.objects.create(**params)
    """
    下面的代码块是为了将模板中的采集项保存到解析表中，通过data字典存入数据库
    """
    # 获取模板ID
    template_id = DeviceTemplates.objects.filter(TemplateName=params["TemplateName"])[0].id
    # 遍历前端传递过来的协议列表，DNC: ["ABB", "KUKA"]  PLC: ["FINS_UDP", "MODBUS_TCP"]
    configs = eval(request.POST["config"])
    # 遍历前端传递过来的协议数据
    for key, values in configs.items():
        # 将每个协议对应的所有需要保存的数据，以对象的形式存储到列表中，进行单协议批量存储，减少数据库的save操作
        data_object_list_plc = []
        data_object_list_dnc = []
        # data用来存储每一条数据，通过data去生成一个ORM对象
        data = {
            "template": template_id,
            "operationFlag": 9
        }
        # 前端传递过来的协议属于key，直接取value即可
        if gateType == "GW_PLC":
            for i in range(len(values)):
                data["name"] = values[i][0]["value"]
                data["type"] = values[i][1]["selectedItem"]
                data["functionCode"] = values[i][2]["selectedItem"]
                data["startAddress"] = values[i][3]["value"]
                data["len"] = values[i][4]["value"]
                data["delayMs"] = values[i][5]["value"]
                data["operationUnit"] = values[i][6]["value"]
                data["dataType"] = ProtocolType.objects.get(TypeList=key)
                data["deviceType"] = key
                if key in protocol_plc_list:
                    data["deviceIP"] = values[i][7]["value"]
                else:
                    data["serialNum"] = values[i][7]["value"]
                    data["deviceID"] = values[i][8]["value"]
                data_object_list_plc.append(ProtocolPLC(**data))
            try:
                ProtocolPLC.objects.bulk_create(data_object_list_plc)
            except Exception as e:
                continue
        else:
            for i in range(0, len(values["normal"])):
                if values["normal"]:
                    data["operationValue"] = values["normal"][i]["name"]
                    data["operationUnit"] = values["normal"][i]["unit"]
                    data["upload"] = values["normal"][i]["upload"]
                    data["Description"] = values["normal"][i]["description"]
                    data["deviceType"] = ProtocolType.objects.get(TypeList=key)
                    data["Description"] = values["normal"][i]["description"]
                    data["configValue"] = values["normal"][i]["param"]
                    data["type"] = \
                        ProtocolParamsDNC.objects.filter(operationValue=data["operationValue"], ProType=key)[0].type
                if values["advanced"]:
                    for k in values["advanced"]:
                        if k["name"] == "IP地址":
                            data["deviceIP"] = k["value"]
                            continue
                    data_object_list_dnc.append(ProtocolDNC(**data))
            try:
                ProtocolDNC.objects.bulk_create(data_object_list_dnc)
            except Exception as e:
                continue
    return restful.ok()


@require_POST
def update_template(request):
    """
    更新协议模板
    :param request:
    :return:
    """
    templateID = request.POST['templateID']
    params = {
        "TemplateName": request.POST['name'],
        "proType": request.POST['type'],
        "TypeList": request.POST['protocol'],
        "operationFlag": request.POST['flag']
    }
    DeviceTemplates.objects.filter(id=templateID).update(**params)
    return restful.ok()


@require_POST
def del_template(request):
    """
    删除协议模板
    :param request:
    :return:
    """
    try:
        idList = [int(i) for i in request.POST["idList"].split(",")]
        # operationFlag更新为0
        DeviceTemplates.objects.filter(id__in=idList).update(operationFlag=0)
        result = restful.ok()
    except Exception as e:
        result = restful.params_error(message="模板删除失败!")
    return restful.ok()


@require_POST
def del_device(request):
    """
    清空网关物影子和相关协议
    :return:
    """
    gateMac = eval(request.POST["gateMac"])
    gateinfos = GateInfoModel.objects.filter(gateMac__in=gateMac)
    for i in gateinfos:
        if i.gateType == "GW_DNC":
            ProtocolDNC.objects.filter(gateMac__in=gateMac).delete()
        else:
            ProtocolPLC.objects.filter(gateMac__in=gateMac).delete()
        Variables.objects.filter(gateMac__in=gateMac).delete()
    DeviceModal.objects.filter(gateMac__in=gateMac).delete()
    return restful.ok()


@require_POST
def update_device(request):
    """
    保存多个物影子，这里只更新物影子的基础信息
    :return:
    """
    data = {}
    deviceInfo = eval(request.POST["deviceInfo"])
    for i in deviceInfo:
        data["gateMac"] = i["gateMac"]
        data["modalName"] = i["deviceName"]
        data["machCode"] = i["deviceCode"]
        data["des"] = i["deviceDesc"]
        device = DeviceModal.objects.filter(machCode=i["deviceCode"])
        if device:
            try:
                DeviceModal.objects.filter(machCode=i["deviceCode"]).update(**data)
            except Exception as e:
                continue
        else:
            DeviceModal.objects.create(**data)
    return restful.ok()


@require_POST
def new_dnc_params(request):
    """
    新建采集项
    :return:
    """
    params = {
        "gateMac": request.POST["gateMac"],
        "type": request.POST["type"],
        "operationFlag": request.POST["operationFlag"],
        "operationUnit": request.POST["operationUnit"],
        "deviceIP": request.POST["deviceIP"],
        "operationValue": request.POST["operationValue"],
        "arrayNum": request.POST["arrayNum"],
        "deviceType_id": 1
    }
    ProtocolDNC.objects.create(**params)
    return restful.ok()


@require_POST
def bulk_import_dnc_params(request):
    """
    批量导入DNC协议采集项目模板
    :return:
    """
    # DNC协议采集项模板文件放置路径
    basic_path = os.path.join(BASE_DIR, "apps", "Tm", "MatterAnalysis", "ProtocolTemp")
    # 设置默认DNC协议采集项模板文件
    file_import_path = os.path.join(BASE_DIR, "apps", "Tm", "MatterAnalysis", "ProtocolTemp", "temp.xlsx")
    # 查找是否有DNC协议采集项模板文件（xlsx/xls）
    for root, dirs, files in os.walk(basic_path):
        for filename in files:
            name, suf = os.path.splitext(filename)
            if (suf == '.xlsx') or (suf == '.xls'):
                file_import_path = os.path.join(root, filename)
    # 确定路径下有模板文件
    if not os.path.exists(file_import_path):
        return restful.params_error("没有找到协议模板文件")
    # 读取excel信息
    workbook = xlrd.open_workbook(file_import_path)
    sheet_list = workbook.sheet_names()
    # 初始化数据
    data = []
    for sheet in sheet_list:
        booksheet = workbook.sheet_by_name(sheet)
        # 获取数据的长宽
        x = booksheet.ncols
        y = booksheet.nrows
        # 获取模板内采集项key
        key = [booksheet.cell_value(0, i) for i in range(x)]
        key.append("协议名称")
        if y:
            for j in range(1, y):
                # 获取采集项信息value
                value = [booksheet.cell_value(j, i) for i in range(x)]
                value.append(sheet)
                info = dict(zip(key, value))
                data.append(info)
    # 将采集项数据与数据库建立映射
    for dt in data:
        basic = {
            "ProType": dt.get("协议名称", ""),
            # operationValue 变量名
            "operationValue": dt.get("Description", ""),
            "Description": dt.get("Title", ""),
            "type": dt.get("变量类型", ""),
            #  采集项单位未配置时默认“-”
            "operationUnit": dt.get("单位", "") if dt.get("单位", "") else "-",
            "configValue": dt.get("configValue", ""),
        }
        # excel导入时数字为float，需转化成int
        if isinstance(basic["Description"], float):
            basic["Description"] = int(basic["Description"])
        try:
            ProtocolParamsDNC.objects.create(**basic)
        except Exception:
            continue
    return restful.ok()


@require_POST
def bulk_export_to_excel(request):
    datas = ProtocolParamsDNC.objects.all()
    data_dict = [j.__dict__ for j in datas]
    sheets = [j for j in set(dt["ProType"] for dt in data_dict)]
    filename = request.POST["filename"]
    column = [u"Description", u"变量类型", u"单位", "Title", u"configValue"]
    colindex = ['Description', 'type', 'operationUnit', 'operationValue', 'configValue']
    # 查看文件格式
    file, Type = os.path.splitext(filename)
    if Type == ".xls":
        export_to_excel_xls(data_dict, filename, sheets, column, colindex)
    elif Type == ".xlsx":
        export_to_excel_xlsx(data_dict, filename, sheets, column, colindex)
    else:
        return restful.method_error("Excel文件名错误！")
    return restful.ok()


def export_to_excel_xls(data, filename, sheets, column, colindex):
    """
    将数据导出到excel
    :param data: 需导出的数据
    :param filename: 文件路径.xls
    :param sheets:
    :param column:
    :param colindex:
    :return:
    """
    wb = xlwt.Workbook(encoding='utf-8')
    # 获取列数
    columns = len(column)
    # 创建格式style
    style = xlwt.XFStyle()
    # 创建font，设置字体
    font = xlwt.Font()
    # 字体格式
    font.name = 'Times New Roman'
    # 将字体font，应用到格式style
    style.font = font
    # 创建alignment，居中
    alignment = xlwt.Alignment()
    # 居中
    alignment.horz = xlwt.Alignment.HORZ_CENTER
    # 应用到格式style
    style.alignment = alignment
    c_style = xlwt.XFStyle()
    c_font = xlwt.Font()
    c_font.name = 'Times New Roman'
    # 字体加粗
    c_font.bold = True
    c_style.font = c_font
    c_style.alignment = alignment
    for sheetname in sheets:
        infos = list(filter(lambda dt: dt["ProType"] == sheetname, data))
        sheet = wb.add_sheet(sheetname, cell_overwrite_ok=True)
        for i in range(columns):
            # 设置列的宽度
            sheet.col(i).width = 5000
            # 插入列名
        for i in range(columns):
            sheet.write(0, i, column[i], c_style)

            # 将数据插入表格
            # 列动行不动
        for i in range(len(infos)):
            for j in range(len(column)):
                sheet.write(i + 1, j, infos[i][colindex[j]], style)
    wb.save(filename)


def export_to_excel_xlsx(data, filename, sheets, column, colindex):
    """
    导出.xlsx为后缀的Excel文件
    :param data:
    :param filename:文件名
    :param sheets:
    :param column:
    :param colindex:
    :return:
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    for sheet in sheets:
        infos = list(filter(lambda dt: dt["ProType"] == sheet, data))
        ws = wb.create_sheet(title=sheet)
        row = len(column)
        # 插入列名
        for i in range(1, row + 1):
            ws.cell(1, i, column[i - 1])
        # 逐条插入数据
        for x in range(1, row + 1):
            for y in range(2, len(infos) + 2):
                ws.cell(y, x, infos[y - 2][colindex[x - 1]])
    # 去掉自动生成的Sheet
    wss = wb["Sheet"]
    wb.remove(wss)

    wb.save(filename)
